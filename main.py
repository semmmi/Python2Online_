# Reading an excel file using Python
import openpyxl as openpyxl
import xlrd as xlrd

from openpyxl import load_workbook

import xlrd

fn = 'Binance3.xlsx'
wb = load_workbook(fn)
ws = wb['listb']

#ws['A8'] = 'ghbdtn'
pr = ws['A8']
print(pr)
#wb.save(fn)
wb.close()
